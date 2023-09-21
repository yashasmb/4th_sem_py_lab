from openpyxl import Workbook  # Import the Workbook class from the openpyxl library.
from openpyxl.styles import Font  # Import the Font class from openpyxl.styles.

wb = Workbook()  # Create a new Excel workbook and assign it to the variable 'wb'.
sheet = wb.active  # Get the active sheet (the first sheet created by default) and assign it to the variable 'sheet'.
sheet.title = "Language"  # Set the title of the active sheet to "Language".

wb.create_sheet(title="Capital")  # Create a new sheet with the title "Capital".

# Define lists for language, state, capital, and code data.
lang = ["Kannada", "Telugu", "Tamil"]
state = ["Karnataka", "Telangana", "Tamil Nadu"]
capital = ["Bengaluru", "Hyderabad", "Chennai"]
code = ['KA', 'TS', 'TN']

# Populate the active sheet with data and add a header row.
sheet.cell(row=1, column=1).value = "State"
sheet.cell(row=1, column=2).value = "Language"
sheet.cell(row=1, column=3).value = "Code"

# Apply bold font to the header row.
ft = Font(bold=True)
for row in sheet["A1:C1"]:
    for cell in row:
        cell.font = ft

# Fill in data for the active sheet.
for i in range(2, 5):
    sheet.cell(row=i, column=1).value = state[i - 2]
    sheet.cell(row=i, column=2).value = lang[i - 2]
    sheet.cell(row=i, column=3).value = code[i - 2]

# Save the workbook to a file named "demo.xlsx".
wb.save("demo.xlsx")

# Switch to the "Capital" sheet.
sheet = wb["Capital"]

# Add headers and apply bold font.
sheet.cell(row=1, column=1).value = "State"
sheet.cell(row=1, column=2).value = "Capital"
sheet.cell(row=1, column=3).value = "Code"

ft = Font(bold=True)
for row in sheet["A1:C1"]:
    for cell in row:
        cell.font = ft

# Fill in data for the "Capital" sheet.
for i in range(2, 5):
    sheet.cell(row=i, column=1).value = state[i - 2]
    sheet.cell(row=i, column=2).value = capital[i - 2]
    sheet.cell(row=i, column=3).value = code[i - 2]

# Save the workbook again to update the "Capital" sheet.
wb.save("demo.xlsx")

# Prompt the user for a state code and search for the corresponding capital in the "Capital" sheet.
srchCode = input("Enter state code for finding capital ")
for i in range(2, 5):
    data = sheet.cell(row=i, column=3).value
    if data == srchCode:
        print("Corresponding capital for code", srchCode, "is", sheet.cell(row=i, column=2).value)

# Switch to the "Language" sheet.
sheet = wb["Language"]

# Prompt the user for a state code and search for the corresponding language in the "Language" sheet.
srchCode = input("Enter state code for finding language ")
for i in range(2, 5):
    data = sheet.cell(row=i, column=3).value
    if data == srchCode:
        print("Corresponding language for code", srchCode, "is", sheet.cell(row=i, column=2).value)

wb.close()  # Close the workbook when done.
